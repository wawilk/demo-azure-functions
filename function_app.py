from io import BytesIO
import azure.functions as func
import datetime
import os
import json
import logging

from certifi import contents
from content_understanding_client import AzureContentUnderstandingClient
from azure.identity import ClientSecretCredential, get_bearer_token_provider
from azure.storage.blob import BlobServiceClient

app = func.FunctionApp()

@app.route(route="perform_ocr", auth_level=func.AuthLevel.ANONYMOUS)
def perform_ocr(req: func.HttpRequest) -> func.HttpResponse:
    """
    Azure Function to perform OCR using Azure Content Understanding
    
    Expected parameters:
    - classifier_id: The classifier ID to use for document processing
    - blob_url: URL of the blob to process
    - storage_account_name: Name of the storage account for results
    """
    logging.info('Python HTTP trigger function processed a request.')

    try:
        # Get parameters from request
        classifier_id = req.params.get('classifier_id')
        blob_url = req.params.get('blob_url')
        storage_account_name = req.params.get('storage_account_name')
        
        # Try to get parameters from request body if not in query string
        if not all([classifier_id, blob_url, storage_account_name]):
            try:
                req_body = req.get_json()
                if req_body:
                    classifier_id = classifier_id or req_body.get('classifier_id')
                    blob_url = blob_url or req_body.get('blob_url')
                    storage_account_name = storage_account_name or req_body.get('storage_account_name')
            except ValueError:
                pass
        
        # Validate required parameters
        if not all([classifier_id, blob_url, storage_account_name]):
            return func.HttpResponse(
                json.dumps({
                    "error": "Missing required parameters",
                    "message": "classifier_id, blob_url, and storage_account_name are required"
                }),
                status_code=400,
                mimetype="application/json"
            )
        
        # Call the OCR function
        result = perform_ocr_processing(classifier_id, blob_url, storage_account_name)
        
        return func.HttpResponse(
            json.dumps({
                "success": True,
                "message": "OCR processing completed successfully",
                "result_blob_name": result.get("blob_name"),
                "container_name": result.get("container_name")
            }),
            status_code=200,
            mimetype="application/json"
        )
        
    except Exception as e:
        logging.error(f"Error in perform_ocr: {str(e)}")
        return func.HttpResponse(
            json.dumps({
                "success": False,
                "error": str(e)
            }),
            status_code=500,
            mimetype="application/json"
        )


def perform_ocr_processing(classifier_id: str, blob_url: str, storage_account_name: str) -> dict:
    """
    Core OCR processing function extracted from the original function
    """
    
    # Get configuration from environment variables
    client_id = os.getenv("AZURE_CLIENT_ID")
    client_secret = os.getenv("AZURE_CLIENT_SECRET")
    tenant_id = os.getenv("AZURE_TENANT_ID")
    endpoint = os.getenv("SERVICE_FOR_CU")
    api_version = os.getenv("SERVICE_API_FOR_CU")
    
    # Validate environment variables
    required_env_vars = [client_id, client_secret, tenant_id, endpoint, api_version]
    if not all(required_env_vars):
        raise ValueError("Missing required environment variables: AZURE_CLIENT_ID, AZURE_CLIENT_SECRET, AZURE_TENANT_ID, SERVICE_FOR_CU, SERVICE_API_FOR_CU")

    # Create the credential
    credential = ClientSecretCredential(tenant_id, client_id, client_secret)
    scope = "https://cognitiveservices.azure.com/.default" 
    token_provider = get_bearer_token_provider(credential, "https://cognitiveservices.azure.com/.default")

    # Initialize the Azure Content Understanding client
    try:
        content_understanding_client = AzureContentUnderstandingClient(
            endpoint=endpoint,
            api_version=api_version,
            token_provider=token_provider,
        )
        logging.info("✅ Content Understanding client initialized successfully!")
    except Exception as e:
        logging.error(f"❌ Failed to initialize client: {e}")
        raise
    
    # Process with classifier
    ocr_result_json = None
    if classifier_id:
        logging.info(f"🔨 Using classifier: {classifier_id}")
        try:
            # Process with enhanced classifier
            logging.info("📄 Processing document with classifier")
            logging.info(f"   Document: {blob_url}")
            logging.info("\n⏳ Processing with classification + field extraction...")

            response = content_understanding_client.begin_classify(classifier_id=classifier_id, file_location=blob_url)
            ocr_result_json = content_understanding_client.poll_result(response, timeout_seconds=920, polling_interval_seconds=25)
            
            logging.info("\n✅ Processing completed!")
            
        except Exception as e:
            logging.error(f"\n❌ Error processing document: {e}")
            raise
    else:
        raise ValueError("⚠️ classifier does not exist.")

    # Upload results to blob storage
    storage_account_url = f"https://{storage_account_name}.blob.core.windows.net"
    blob_service_client = BlobServiceClient(account_url=storage_account_url, credential=credential)
    container_name = "enhanced-results"
    
    # Generate result file name
    original_file_name = os.path.basename(blob_url)
    ocr_result_blob_name = f"{original_file_name}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.json"

    # Convert the JSON object to a string
    json_data = json.dumps(ocr_result_json)

    # get the ocr_result_json JSON from the blob storage
    blob_client = blob_service_client.get_blob_client(container=container_name, blob=ocr_result_blob_name)
    blob_client.upload_blob(json_data, overwrite=True)
    logging.info(f"Blob '{ocr_result_blob_name}' uploaded successfully to container '{container_name}'.")
    
    return {
        "blob_name": ocr_result_blob_name,
        "container_name": container_name
    }

@app.route(route="create_excel", auth_level=func.AuthLevel.ANONYMOUS)
def create_excel(req: func.HttpRequest) -> func.HttpResponse:
    """
    Azure Function to create an excel spreadsheet report from content 
    in the OCR results blob (same storage account, different container)
    Expected parameters:
    - ocr_result_blob_name: The name of the JSON blob to process
    - storage_account_name: Name of the storage account for results
    """
    try:
        # **************************************************************
        # the format for the blob names is as follows
        # orig PDF name + timestamp + .json
        # the OCR result and the excel result will both use the
        # same base blob file name only the extension is different
        # examples:
        # original PDF name = sample claim submission.pdf
        # ocr_result_blob_name = sample claim submission.pdf_20251016_111305.json
        # excel_result_blob_name = sample claim submission.pdf_20251016_111305.xlsx
        # **************************************************************

        # Get parameters from request
        ocr_result_blob_name = req.params.get('ocr_result_blob_name')

        # storage account is same for source and target
        storage_account_name = req.params.get('storage_account_name')

        # Get configuration from environment variables
        client_id = os.getenv("AZURE_CLIENT_ID")
        client_secret = os.getenv("AZURE_CLIENT_SECRET")
        tenant_id = os.getenv("AZURE_TENANT_ID")
        credential = ClientSecretCredential(tenant_id, client_id, client_secret)

        # this is the url of the storage account
        ocr_results_container_name = "enhanced-results"
        storage_account_url = f"https://{storage_account_name}.blob.core.windows.net"

        ocr_result_blob_url = storage_account_url + "/" + ocr_results_container_name + "/" + ocr_result_blob_name

        # Try to get parameters from request body if not in query string
        if not all([ocr_result_blob_name, storage_account_name]):
            try:
                req_body = req.get_json()
                if req_body:
                    ocr_result_blob_name = ocr_result_blob_name or req_body.get('ocr_result_blob_name')
                    storage_account_name = storage_account_name or req_body.get('storage_account_name')
            except ValueError:
                pass
        
        # Validate required parameters
        if not all([ocr_result_blob_name, storage_account_name]):
            return func.HttpResponse(
                json.dumps({
                    "error": "Missing required parameters",
                    "message": "ocr_result_blob_name and storage_account_name are required",
                    "received": {
                        "req": req                    
                    }
                }),
                status_code=400,
                mimetype="application/json"
            )
        # Get the enhanced result from the blob url and the enhance_results container
        
        ocr_blob_service_client = BlobServiceClient(account_url=storage_account_url, credential=credential)

        # Get the BlobClient for the specific blob
        blob_client = ocr_blob_service_client.get_blob_client(container=ocr_results_container_name, blob=ocr_result_blob_name)

        # Download the blob content into a variable and parse as JSON
        blob_content = blob_client.download_blob().readall()
        ocr_result_json = json.loads(blob_content.decode('utf-8'))
        
        print("starting excel")

        # Call the create_excel function to produce the report content
        excel_file_content = produce_excel_report(ocr_result_json, storage_account_name, ocr_result_blob_name)

        # return success to the caller
        return func.HttpResponse(
            json.dumps({
                "success": True,
                "message": "Excel File blob created successfully",
                "result_blob_name": excel_file_content.get("excel_blob_name"),
                "container_name": excel_file_content.get("container_name")
            }),
            status_code=200,
            mimetype="application/json"
        )
    
    # error occurred
    except Exception as e:
        logging.error(f"Error in create_excel: {str(e)}")
        return func.HttpResponse(
            json.dumps({
                "success": False,
                "error": str(e),
                "storage_account_name": storage_account_name,
                "ocr_result_blob_name": ocr_result_blob_name
            }),
            status_code=500,
            mimetype="application/json"
        )
    
def determine_original_pdfs_name(blob_url):
    """
    Extract the original PDF file name from the blob URL.
    An enhanced result file name will be like this: sample claim submission.pdf_20251016_111305.json
    The excel name should be like this: sample claim submission.pdf_20251016_111305.xlsx
    """
    import os
    return os.path.basename(blob_url).split('.')[0]
    
def produce_excel_report(ocr_result_json, storage_account_name, ocr_blob_name):
    """
    Create an Excel report from the ocr_result_json data with patient info, 
    document listings, and collapsible expense rows.
    and writes the report to a blob file in Azure storage
    Returns the blob name and container name of the uploaded excel file
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from datetime import datetime
    import os

    # Get configuration from environment variables
    client_id = os.getenv("AZURE_CLIENT_ID")
    client_secret = os.getenv("AZURE_CLIENT_SECRET")
    tenant_id = os.getenv("AZURE_TENANT_ID")

    # the credential should be a service account
    credential = ClientSecretCredential(tenant_id, client_id, client_secret)

    try:
        # data is the output from the OCR step
        print("📊 Creating Excel Report")
        print("=" * 50)
        
        ocr_result_data = ocr_result_json.get("result", {})
        ocr_contents = ocr_result_data.get("contents", [])

        if not ocr_contents:
            print("❌ No data to export")
            return None
        
        # Create workbook and worksheet
        # the output from this is the wb object
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Claims Analysis Report"
        
        # Define styles
        header_font = Font(bold=True, size=12)
        subheader_font = Font(bold=True, size=10)
        normal_font = Font(size=10)
        patient_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        doc_fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
        expense_fill = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")
        
        # Borders
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        current_row = 1
        
        # Extract patient information from claim form (first document that has patient info)
        patient_info = {}
        for content in ocr_contents:
            fields = content.get('fields', {})
            if any(field in fields for field in ['Patient_First_Name', 'Patient_Last_Name', 'DOB', 'Gender', 'Policy_Number']):
                patient_info = {
                    'first_name': fields.get('Patient_First_Name', {}).get('valueString', ''),
                    'last_name': fields.get('Patient_Last_Name', {}).get('valueString', ''),
                    'dob': fields.get('DOB', {}).get('valueString', ''),
                    'gender': fields.get('Gender', {}).get('valueString', ''),
                    'policy_number': fields.get('Policy_Number', {}).get('valueString', '')
                }
                break
        
        # Patient Information Section
        ws.merge_cells(f'A{current_row}:I{current_row}')
        ws[f'A{current_row}'] = "PATIENT INFORMATION"
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
        current_row += 1
        
        # Patient Name row
        patient_name = f"{patient_info.get('first_name', '')} {patient_info.get('last_name', '')}".strip()
        ws[f'A{current_row}'] = "Patient Name:"
        ws[f'B{current_row}'] = patient_name
        ws[f'A{current_row}'].font = subheader_font
        ws[f'B{current_row}'].font = normal_font
        current_row += 1
        
        # Patient details headers
        ws[f'A{current_row}'] = "DOB"
        ws[f'C{current_row}'] = "Gender"
        ws[f'E{current_row}'] = "Policy Number"
        for col in ['A', 'C', 'E']:
            ws[f'{col}{current_row}'].font = subheader_font
        current_row += 1
        
        # Patient details values
        ws[f'A{current_row}'] = patient_info.get('dob', '')
        ws[f'C{current_row}'] = patient_info.get('gender', '')
        ws[f'E{current_row}'] = patient_info.get('policy_number', '')
        for col in ['A', 'C', 'E']:
            ws[f'{col}{current_row}'].font = normal_font
        
        # Apply patient info styling
        for row in range(1, current_row + 1):
            for col in range(1, 11):  # A to I
                cell = ws.cell(row=row, column=col)
                cell.fill = patient_fill
                cell.border = thin_border
        
        current_row += 3  # Add spacing
        
        # Document Section Header
        ws.merge_cells(f'A{current_row}:I{current_row}')
        ws[f'A{current_row}'] = "DOCUMENTS FOUND IN BUNDLE"
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
        current_row += 1
        
        # Document headers
        doc_headers = ['Document #', 'Title', 'Starting Page', 'Ending Page', 'Number of Pages']
        for i, header in enumerate(doc_headers, 1):
            ws.cell(row=current_row, column=i, value=header)
            ws.cell(row=current_row, column=i).font = subheader_font
            ws.cell(row=current_row, column=i).fill = doc_fill
            ws.cell(row=current_row, column=i).border = thin_border

        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center')
        current_row += 1
        
        # Process each document
        for doc_num, content in enumerate(ocr_contents, 1):
            doc_start_row = current_row
            
            # Document basic info
            category = content.get('category', 'Unknown')
            start_page = content.get('startPageNumber', '?')
            end_page = content.get('endPageNumber', '?')
            
            if start_page != '?' and end_page != '?':
                num_pages = end_page - start_page + 1
            else:
                num_pages = '?'
            
            # Get document title
            fields = content.get('fields', {})
            title = "N/A"
            if 'title_on_first_page_of_document' in fields:
                title = fields['title_on_first_page_of_document'].get('valueString', 'N/A')
            
            # Write document row
            doc_data = [doc_num, title, start_page, end_page, num_pages]
            for i, value in enumerate(doc_data, 1):
                ws.cell(row=current_row, column=i, value=value)
                ws.cell(row=current_row, column=i).font = normal_font
                ws.cell(row=current_row, column=i).fill = doc_fill
                ws.cell(row=current_row, column=i).border = thin_border
            
            # Extend document row background color to column k (10) to line up with expense columns
            for col in range(6, 11):  # Columns F through k
                ws.cell(row=current_row, column=col).fill = doc_fill
                ws.cell(row=current_row, column=col).border = thin_border
            
            # center column A
            ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center')

            current_row += 1
            
            # Check for expenses
            if 'Expenses' in fields:
                expenses = fields['Expenses'].get('valueArray', [])
                if expenses:
                    # Expense headers (starting from column B, leaving A empty)
                    expense_headers = [
                        'Expense Amount', 'Expense Description', 'Date', 'CPT Code',
                        'ICD Code', 'Expense Type', 'Surgeon/Provider', 'Ref Page', 'Drug Name'
                    ]
                    
                    for i, header in enumerate(expense_headers, 2):  # Start from column B (2)
                        ws.cell(row=current_row, column=i, value=header)
                        ws.cell(row=current_row, column=i).font = subheader_font
                        ws.cell(row=current_row, column=i).fill = expense_fill
                        ws.cell(row=current_row, column=i).border = thin_border
                    
                    expense_start_row = current_row
                    current_row += 1
                    
                    # Process expenses
                    for expense in expenses:
                        expense_obj = expense.get('valueObject', {})
                        
                        # Extract expense data
                        expense_data = []
                        
                        # Expense Amount
                        amount_field = expense_obj.get('Expense_Amount', {})
                        if amount_field.get('type') == 'number':
                            amount = amount_field.get('valueNumber', 0)
                            expense_data.append(f"${amount:.2f}")
                        else:
                            expense_data.append('N/A')
                        
                        # Description
                        desc_field = expense_obj.get('Expense_Description', {})
                        expense_data.append(desc_field.get('valueString', 'N/A'))
                        
                        # Date
                        date_field = expense_obj.get('Date', {})
                        if date_field.get('type') == 'date':
                            expense_data.append(date_field.get('valueDate', 'N/A'))
                        else:
                            expense_data.append(date_field.get('valueString', 'N/A'))
                        
                        # CPT Code
                        cpt_field = expense_obj.get('CPT_Code', {})
                        expense_data.append(cpt_field.get('valueString', 'N/A'))
                        
                        # ICD Code
                        icd_field = expense_obj.get('ICD_Code', {})
                        expense_data.append(icd_field.get('valueString', 'N/A'))
                        
                        # Expense Type
                        type_field = expense_obj.get('Expense_Type', {})
                        expense_data.append(type_field.get('valueString', 'N/A'))
                        
                        # Surgeon/Provider
                        surgeon_field = expense_obj.get('Surgeon_Name_or_Provider', {})
                        expense_data.append(surgeon_field.get('valueString', 'N/A'))
                        
                        # Ref Page
                        ref_field = expense_obj.get('Ref_Page', {})
                        if ref_field.get('type') == 'number':
                            ref_page = ref_field.get('valueNumber', 0)
                            # Adjust page number relative to document start
                            adjusted_page = int(ref_page) + start_page - 1 if start_page != '?' else ref_page
                            expense_data.append(adjusted_page)
                        else:
                            expense_data.append('N/A')
                        
                        # Drug Name
                        drug_field = expense_obj.get('Drug_Name', {})
                        expense_data.append(drug_field.get('valueString', 'N/A'))
                        
                        # Write expense row (starting from column B, leaving A empty)
                        for i, value in enumerate(expense_data, 2):  # Start from column B (2)
                            ws.cell(row=current_row, column=i, value=value)
                            ws.cell(row=current_row, column=i).font = normal_font
                            ws.cell(row=current_row, column=i).fill = expense_fill
                            ws.cell(row=current_row, column=i).border = thin_border
                        
                        current_row += 1
                    
                    # Create collapsible group for expenses (including header and expense rows)
                    if current_row > expense_start_row + 1:
                        ws.row_dimensions.group(expense_start_row, current_row - 1, hidden=True)
            
            current_row += 1  # Add spacing between documents
        
        # Auto-size columns (simplified approach to avoid merged cell issues)
        for col_num in range(1, 11):  # Columns A through J
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(col_num)
            
            # Iterate through all rows for this column
            for row_num in range(1, current_row):
                cell = ws.cell(row=row_num, column=col_num)
                
                # Skip merged cells to avoid conflicts
                is_merged = False
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        is_merged = True
                        break
                
                # Only calculate length for non-merged cells with values
                if not is_merged and cell.value:
                    cell_length = len(str(cell.value))
                    max_length = max(max_length, cell_length)
            
            # Set column width (cap at 50 characters for readability)
            adjusted_width = min(max_length + 2, 50) if max_length > 0 else 15
            ws.column_dimensions[column_letter].width = adjusted_width
            # Save the workbook
            output = BytesIO()
            wb.save(output)
        
    except Exception as e:
            print(f"❌ Error creating Excel report: {e}")
            return None

    try:
        # ***************************************************
        # We are uploading the excel to blob storage
        # ***************************************************
        # we want to upload to the excel-result container
        excel_container_name = "excel-result"

        # Generate excel file name from the ocr blob file name
        # just change the extension to xlsx
        excel_blob_name = ocr_blob_name.rsplit('.', 1)[0] + ".xlsx"

        # same storage account url as before
        storage_account_url = f"https://{storage_account_name}.blob.core.windows.net"

        # The excel blob url includes the excel_blob name
        excel_blob_url = storage_account_url + "/" + excel_container_name + "/" + excel_blob_name
        excel_blob_service_client = BlobServiceClient(account_url=storage_account_url, credential=credential)
        # use the blob service client to get a blob client
        excel_blob_client = excel_blob_service_client.get_blob_client(container=excel_container_name, blob=excel_blob_name)

        # ***************************************************
        # Upload the excel report content file to the excel
        # container blob storage.
        # output.getvalue() contains the contents of the excel file
        # ***************************************************
        excel_blob_client.upload_blob(output.getvalue(), overwrite=True)

        # log it
        logging.info(f"Blob '{excel_blob_name}' uploaded successfully to excel container '{excel_container_name}'.")

        # return the excel blob name and container name
        return {
            "blob_name": excel_blob_name,
            "container_name": excel_container_name
        }
    except Exception as e:
        logging.info(f"❌ Error creating Excel report: {e}")
        logging.info(f" The connection string was {storage_account_url}")
        logging.info(f" The storage_account_name was {storage_account_name}")
        logging.info(f" The target excel container_name was {excel_container_name}")
        logging.info(f" The blob name was {excel_blob_name}")

        # if there was an error, return what we know
        if excel_blob_name is None:
            excel_blob_name = "unknown_due_to_error"
        if excel_container_name is None:
            excel_container_name = "unknown_due_to_error"
        return {
            "blob_name": excel_blob_name,
            "container_name": excel_container_name
        }

    # to test the create excel
    # http://localhost:7071/api/create_excel?ocr_result_blob_name=%22https://wwawilkdemostow.blob.core.windows.net/incoming-docs/sample%20claim%20submission.pdf_20251016_111305.json%22&storage_account_name=%22wwawilkdemostow%22

@app.route(route="clean_up", auth_level=func.AuthLevel.ANONYMOUS)
def clean_up(req: func.HttpRequest) -> func.HttpResponse:
    """
    Azure Function to perform clean up of blob in containers
    This function should move the blob that is in the incoming-docs container
    to the processed-docs container or delete it.
    
    Expected parameters:
    - incoming_docs_blob_name: URL of the blob to clean up
    - storage_account_name: Name of the storage account for results
    """
    # Get parameters from request
    incoming_docs_blob_name = req.params.get('incoming_docs_blob_name')

    # storage account is same for source and target
    storage_account_name = req.params.get('storage_account_name')

    # Get configuration from environment variables
    client_id = os.getenv("AZURE_CLIENT_ID")
    client_secret = os.getenv("AZURE_CLIENT_SECRET")
    tenant_id = os.getenv("AZURE_TENANT_ID")
    credential = ClientSecretCredential(tenant_id, client_id, client_secret)

    # this is the url of the storage account
    incoming_docs_container_name = "incoming-docs"
    storage_account_url = f"https://{storage_account_name}.blob.core.windows.net"

    incoming_docs_blob_url = storage_account_url + "/" + incoming_docs_container_name + "/" + incoming_docs_blob_name

    # Try to get parameters from request body if not in query string
    if not all([incoming_docs_blob_name, storage_account_name]):
        try:
            req_body = req.get_json()
            if req_body:
                incoming_docs_blob_name = incoming_docs_blob_name or req_body.get('incoming_docs_blob_name')
                storage_account_name = storage_account_name or req_body.get('storage_account_name')
        except ValueError:
            pass
    
    # Validate required parameters
    if not all([incoming_docs_blob_name, storage_account_name]):
        return func.HttpResponse(
            json.dumps({
                "error": "Missing required parameters",
                "message": "incoming_docs_blob_name and storage_account_name are required",
                "received": {
                    "req": req                    
                }
            }),
            status_code=400,
            mimetype="application/json"
        )
    # we are getting the PDF from incoming docs and putting into the processed-docs container
    # before we delete it from incoming docs
    try:
        logging.info(f"🔄 Cleaning up blob: {incoming_docs_blob_name} from {incoming_docs_container_name} container")
        incoming_docs_blob_service_client = BlobServiceClient(account_url=storage_account_url, credential=credential)
        processed_docs_blob_service_client = BlobServiceClient(account_url=storage_account_url, credential=credential)
        
        # Get the BlobClient for the specific blob
        incoming_docs_blob_client = incoming_docs_blob_service_client.get_blob_client(container=incoming_docs_container_name, blob=incoming_docs_blob_name)

        # get the blob from the client above incoming_docs_blob_client
        incoming_docs_blob = incoming_docs_blob_client.download_blob().readall()

        # write it out first before deleting
        processed_docs_blob_client = processed_docs_blob_service_client.get_blob_client(container="processed-docs", blob=incoming_docs_blob_name)
        processed_docs_blob_client.upload_blob(incoming_docs_blob, overwrite=True)
        logging.info(f"Blob '{incoming_docs_blob_name}' uploaded successfully to container 'processed-docs'.")

        # Delete the blob from incoming -docs container
        incoming_docs_blob_client.delete_blob()
        logging.info(f"Blob '{incoming_docs_blob_name}' deleted successfully from container {incoming_docs_container_name}.")

        return func.HttpResponse(
            json.dumps({
                "success": True,
                "message": f"Blob '{incoming_docs_blob_name}' cleaned up successfully from container {incoming_docs_container_name}."
            }),
            status_code=200,
            mimetype="application/json"
        )
    except Exception as e:
        logging.error(f"Error in clean_up: {str(e)}")
        return func.HttpResponse(
            json.dumps({
                "success": False,
                "error": str(e)
            }),
            status_code=500,
            mimetype="application/json"
        )
    # we can test with curl -G "https://fat-1488137190.azurewebsites.net/api/clean_up" --data-urlencode "incoming_docs_blob_name=sample claim submission.pdf"  --data-urlencode "storage_account_name=wwawilkdemostow"

@app.route(route="parse_ocr", auth_level=func.AuthLevel.ANONYMOUS)
def parse_ocr(req: func.HttpRequest) -> func.HttpResponse:
    """
    Azure Function to create a summary report from content 
    in the OCR results blob (same storage account, different container)
    Expected parameters:
    - ocr_result_blob_name: The name of the JSON blob to process
    - storage_account_name: Name of the storage account for results
    """
    try:
        # **************************************************************
        # the format for the blob names is as follows
        # orig PDF name + timestamp + .json
        # the OCR result and the excel result will both use the
        # same base blob file name only the extension is different
        # examples:
        # original PDF name = sample claim submission.pdf
        # ocr_result_blob_name = sample claim submission.pdf_20251016_111305.json
        # excel_result_blob_name = sample claim submission.pdf_20251016_111305.xlsx
        # summary_report_blob_name = sample claim submission.pdf_20251016_111305_summary.txt
        # **************************************************************

        # Get parameters from request
        ocr_result_blob_name = req.params.get('ocr_result_blob_name')

        # storage account is same for source and target
        storage_account_name = req.params.get('storage_account_name')

        # Get configuration from environment variables
        client_id = os.getenv("AZURE_CLIENT_ID")
        client_secret = os.getenv("AZURE_CLIENT_SECRET")
        tenant_id = os.getenv("AZURE_TENANT_ID")
        credential = ClientSecretCredential(tenant_id, client_id, client_secret)

        # this is the url of the storage account
        ocr_results_container_name = "enhanced-results"
        storage_account_url = f"https://{storage_account_name}.blob.core.windows.net"

        ocr_result_blob_url = storage_account_url + "/" + ocr_results_container_name + "/" + ocr_result_blob_name

        # Try to get parameters from request body if not in query string
        if not all([ocr_result_blob_name, storage_account_name]):
            try:
                req_body = req.get_json()
                if req_body:
                    ocr_result_blob_name = ocr_result_blob_name or req_body.get('ocr_result_blob_name')
                    storage_account_name = storage_account_name or req_body.get('storage_account_name')
            except ValueError:
                pass
        
        # Validate required parameters
        if not all([ocr_result_blob_name, storage_account_name]):
            return func.HttpResponse(
                json.dumps({
                    "error": "Missing required parameters",
                    "message": "ocr_result_blob_name and storage_account_name are required",
                    "received": {
                        "req": req                    
                    }
                }),
                status_code=400,
                mimetype="application/json"
            )
        # Get the enhanced result from the blob url and the enhance_results container
        
        ocr_blob_service_client = BlobServiceClient(account_url=storage_account_url, credential=credential)

        # Get the BlobClient for the specific blob
        blob_client = ocr_blob_service_client.get_blob_client(container=ocr_results_container_name, blob=ocr_result_blob_name)

        # Download the blob content into a variable and parse as JSON
        blob_content = blob_client.download_blob().readall()
        ocr_result_json = json.loads(blob_content.decode('utf-8'))
        
        print("starting summary")

        # Call the create_summary function to produce the report content
        summary_report_blob_name, summary_container_name = produce_summary_report(ocr_result_json, storage_account_name, ocr_result_blob_name, credential)

        # return success to the caller
        return func.HttpResponse(
            json.dumps({
                "success": True,
                "message": "Summary File blob created successfully",
                "summary_report_blob_name": summary_report_blob_name,
                "summary_container_name": summary_container_name
            }),
            status_code=200,
            mimetype="application/json"
        )
    
    # error occurred
    except Exception as e:
        logging.error(f"Error in create_summary: {str(e)}")
        return func.HttpResponse(
            json.dumps({
                "success": False,
                "error": str(e),
                "storage_account_name": storage_account_name,
                "summary_report_blob_name": summary_report_blob_name
            }),
            status_code=500,
            mimetype="application/json"
        )
def produce_summary_report(ocr_result_json, storage_account_name, ocr_blob_name, credential):   
    """
    Display a concise summary of the document analysis results.
    """
    try:
        data = ocr_result_json
        # all of the print statement need to write to a variable that will be the report content
        report_content = []
        report_content.append("📊 DOCUMENT BUNDLE SUMMARY")
        report_content.append("=" * 50)

        result_data = data.get("result", {})
        ocr_contents = result_data.get("contents", [])
        if ocr_contents:  
            last_end_page = ocr_contents[-1].get('endPageNumber', '?')  
        else:  
            last_end_page = '?'

        report_content.append(f"Total documents found: {len(ocr_contents)}")
        report_content.append(f"Total pages in bundle: {last_end_page}")

        # Summary table
        report_content.append("\n📋 Document Summary:")
        report_content.append("-" * 80)
        report_content.append(f"{'#':<3} {'Document Type':<50} {'Pages':<8} {'Fields':<8}")
        report_content.append("-" * 80)
        
        total_expenses = 0
        for i, content in enumerate(ocr_contents, 1):
            category = content.get('category', 'Unknown')
            start_page = content.get('startPageNumber', '?')
            end_page = content.get('endPageNumber', '?')
            
            if start_page != '?' and end_page != '?':
                page_range = f"{start_page}-{end_page}"
                num_pages = end_page - start_page + 1
            else:
                page_range = '?'
                num_pages = '?'
            
            fields = content.get('fields', {})
            field_count = len(fields)
            
            # Count expenses
            if 'Expenses' in fields:
                expenses = fields['Expenses'].get('valueArray', [])
                expense_count = len(expenses)
                total_expenses += expense_count
                field_info = f"{field_count} (+{expense_count} expenses)"
            else:
                field_info = str(field_count)

            report_content.append(f"{i:<3} {category:<50} {page_range:<8} {field_info:<8}")

        report_content.append("-" * 80)
        report_content.append(f"\n💰 Total expenses found across all documents: {total_expenses}")

        # Show which documents have patient info vs expenses
        report_content.append(f"\n📝 Field Distribution:")
        report_content.append(f"   • Insurance Claim Form: Patient information fields")
        report_content.append(f"   • Billing Statements: Expense details + document titles")
        report_content.append(f"   • Other Documents: Document titles only")

        # Join all report content into a single string
        # return "\n".join(report_content)

    except NameError:
        print("❌ Error: enhanced_result variable not found. Run the document processing cell first.")
        return None

    try:
        # ***************************************************
        # We are uploading the excel to blob storage
        # ***************************************************
        # we want to upload to the summary-reports container
        summary_container_name = "summary-reports"

        # Generate summary report file name from the ocr blob file name
        # just change the extension to txt
        summary_report_blob_name = ocr_blob_name.rsplit('.', 1)[0] + ".txt"

        # same storage account url as before
        storage_account_url = f"https://{storage_account_name}.blob.core.windows.net"

        # The summary report blob url includes the summary_report_blob_name
        summary_report_blob_url = storage_account_url + "/" + summary_container_name + "/" + summary_report_blob_name
        summary_report_blob_service_client = BlobServiceClient(account_url=storage_account_url, credential=credential)
        # use the blob service client to get a blob client
        summary_report_blob_client = summary_report_blob_service_client.get_blob_client(container=summary_container_name, blob=summary_report_blob_name)

        # ***************************************************
        # Upload the txt report content file to the summary
        # report container blob storage
        # ***************************************************
        # combine rows
        report_content = "\n".join(report_content)

        #upload as a string
        summary_report_blob_client.upload_blob(report_content, overwrite=True)

        # log it
        logging.info(f"Blob '{summary_report_blob_name}' uploaded successfully to summary container '{summary_container_name}'.")

        # return the summary report blob name and container name
        return {
            "summary_report_blob_name": summary_report_blob_name,
            "summary_container_name": summary_container_name
        }
    except Exception as e:
        logging.info(f"❌ Error creating summary report: {e}")
        logging.info(f" The connection string was {storage_account_url}")
        logging.info(f" The storage_account_name was {storage_account_name}")
        logging.info(f" The target summary container_name was {summary_container_name}")
        logging.info(f" The blob name was {summary_report_blob_name}")

        # if there was an error, return what we know
        if summary_report_blob_name is None:
            summary_report_blob_name = "unknown_due_to_error"
        if summary_container_name is None:
            summary_container_name = "unknown_due_to_error"
        return {
            "summary_report_blob_name": summary_report_blob_name,
            "summary_container_name": summary_container_name
        }
    #test with curl -G "https://fat-1488137190.azurewebsites.net/api/parse_ocr" --data-urlencode "ocr_result_blob_name=sample claim submission.pdf_20251016_111305.json"  --data-urlencode "storage_account_name=wwawilkdemostow"